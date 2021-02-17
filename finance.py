import sys
import os
import string
from enum import Enum
from openpyxl import load_workbook
from openpyxl import Workbook

alphabet_uppercase = string.ascii_uppercase

class Payment:
    def __init__(self, datetime, comment):
        self.datetime = datetime
        self.comment = comment
        self.costs_per_letter = {}

    def set_cost(self, letter, cost):
        if type(letter) != type('a'):
            aaa
        self.costs_per_letter[letter] = cost

    def __eq__(self, other):
        cost_equals = True
        for cost_type in self.costs_per_letter.keys():
            if cost_type not in other.costs_per_letter.keys() or self.costs_per_letter[cost_type] != other.costs_per_letter[cost_type]:
                cost_equals == False
        return self.datetime == other.datetime and self.comment == other.comment and cost_equals

    def __lt__(self, other):
        y_self = int(str.split(self.datetime, '-')[0])
        m_self = int(str.split(self.datetime, '-')[1])
        d_self = int(str.split(self.datetime, '-')[2])

        y_other = int(str.split(other.datetime, '-')[0])
        m_other = int(str.split(other.datetime, '-')[1])
        d_other = int(str.split(other.datetime, '-')[2])

        if y_self < y_other:
            return True
        elif y_self == y_other and m_self < m_other:
            return True
        elif y_self == y_other and m_self == m_other and d_self < d_other:
            return True
        return False

    def __gt__(self, other):
        y_self = int(str.split(self.datetime, '-')[0])
        m_self = int(str.split(self.datetime, '-')[1])
        d_self = int(str.split(self.datetime, '-')[2])

        y_other = int(str.split(other.datetime, '-')[0])
        m_other = int(str.split(other.datetime, '-')[1])
        d_other = int(str.split(other.datetime, '-')[2])

        if y_self > y_other:
            return True
        elif y_self == y_other and m_self > m_other:
            return True
        elif y_self == y_other and m_self == m_other and d_self > d_other:
            return True
        return False


    def similar_to(self, other):
        total_cost_1 = 0
        for letter in self.costs_per_letter.keys():
            total_cost_1 += abs(self.costs_per_letter[letter])

        total_cost_2 = 0
        for letter in other.costs_per_letter.keys():
            total_cost_2 += abs(other.costs_per_letter[letter])

        return self.datetime == other.datetime and self.comment == other.comment and total_cost_1 == total_cost_2

class CostTypeRule:
    def __init__(self, from_field, to_field):
        self.from_field = from_field
        self.to_field = to_field

def generate_header(sheet):
    sheet["C2"].value = "Inkomster"
    sheet["D2"].value = "Konton"
    sheet["G2"].value = "Utgifter"

    sheet["C3"].value = "Rutin"
    sheet["D3"].value = "E-sparkonto"
    sheet["E3"].value = "Investerarkonto"
    sheet["F3"].value = "Kortkonto"
    sheet["G3"].value = "Rutin"
    sheet["H3"].value = "Mat"
    sheet["I3"].value = "Hushåll"
    sheet["J3"].value = "Äta ute"
    sheet["K3"].value = "Hobby/Nöje"
    sheet["L3"].value = "Resa"
    sheet["M3"].value = "Kläder"
    sheet["N3"].value = "Övrigt"
    sheet["O3"].value = "Kommentar"

    sheet["A4"].value = "Budgetering"
    sheet["D4"].value = 0
    sheet["E4"].value = 0
    sheet["G4"].value = 0
    sheet["H4"].value = 0
    sheet["I4"].value = 0
    sheet["J4"].value = 0
    sheet["K4"].value = 0
    sheet["L4"].value = 0
    sheet["M4"].value = 0
    sheet["N4"].value = 0

    sheet["A5"].value = "Ingående balans"
    sheet["D5"].value = initial_balances["Transaktioner e-sparkonto"]
    sheet["E5"].value = initial_balances["Transaktioner Aktiekonto"]
    sheet["F5"].value = initial_balances["Transaktioner Privatkonto"]

def generate_footer(sheet, offset_y):
    sheet['B'+str(offset_y)]="Differens"
    for c in alphabet_uppercase[2:14:1] : 
        sheet[c+str(offset_y)]="=SUM("+c+"7:"+c+str(offset_y-2)+")"

    sheet['B'+str(offset_y+1)]="Utgående saldo"
    sheet['D'+str(offset_y+1)]="=SUM(D5,D"+str(offset_y)+")"
    sheet['E'+str(offset_y+1)]="=SUM(E5,E"+str(offset_y)+")"
    sheet['F'+str(offset_y+1)]="=SUM(F5,F"+str(offset_y)+")"

    sheet["D"+str(offset_y+3)].value = "Sparkonto"
    sheet["E"+str(offset_y+3)].value = "Aktiekonto"

    sheet["G"+str(offset_y+3)].value = "Rutin"
    sheet["H"+str(offset_y+3)].value = "Mat"
    sheet["I"+str(offset_y+3)].value = "Hushåll"
    sheet["J"+str(offset_y+3)].value = "Hobby"
    sheet["K"+str(offset_y+3)].value = "Nöje"
    sheet["L"+str(offset_y+3)].value = "Resa"
    sheet["M"+str(offset_y+3)].value = "Kläder"
    sheet["N"+str(offset_y+3)].value = "Övrigt"
    sheet["O"+str(offset_y+3)].value = "Kommentar"

    sheet["B"+str(offset_y+4)].value = "Budgetdifferens"
    sheet["B"+str(offset_y+5)].value = "Budgetdifferens, carryover"
    sheet["B"+str(offset_y+6)].value = "Budgetdifferens, manuella ändringar till nästa månad"
    sheet["B"+str(offset_y+7)].value = "Budget, ackumulerad nästa månad"

    sheet['D'+str(offset_y+4)]="=SUM(D4,D"+str(offset_y)+")"
    sheet['E'+str(offset_y+4)]="=SUM(E4,E"+str(offset_y)+")"

    for c in alphabet_uppercase[6:14:1] : 
        sheet[c+str(offset_y+4)]="=SUM("+c+"4,-"+c+str(offset_y)+")"
        sheet[c+str(offset_y+5)]=0
        sheet[c+str(offset_y+6)]=0
        sheet[c+str(offset_y+7)]="=SUM("+c+str(offset_y+4)+","+c+str(offset_y+5)+","+c+str(offset_y+6)+")"

    sheet["D"+str(offset_y+7)]="=SUM(D"+str(offset_y+4)+",D"+str(offset_y+5)+",D"+str(offset_y+6)+")"
    sheet["E"+str(offset_y+7)]="=SUM(E"+str(offset_y+4)+",E"+str(offset_y+5)+",E"+str(offset_y+6)+")"

    sheet["B"+str(offset_y+9)]="Kontroll"
    sheet["B"+str(offset_y+10)]="Utgående saldo"
    sheet["B"+str(offset_y+11)]="Ingående saldo"
    sheet["B"+str(offset_y+12)]="Differens"

    sheet["C"+str(offset_y+9)]="=SUM(C"+str(offset_y)+":N"+str(offset_y)+")"
    sheet["C"+str(offset_y+10)]="=SUM(D"+str(offset_y+1)+":N"+str(offset_y+1)+")"
    sheet["C"+str(offset_y+11)]="=SUM(D6:F6)"
    sheet["C"+str(offset_y+12)]="=C"+str(offset_y+10)+"-C"+str(offset_y+11)

def put_payments(sheet, payments):
    i=7
    for payment in payments:
        sheet['B' + str(i)].value = payment.datetime
        sheet['O' + str(i)].value = payment.comment
        for letter in payment.costs_per_letter.keys():
            cost = payment.costs_per_letter[letter]
            sheet[letter + str(i)].value = -cost
        i+=1

# Config

class CostType(Enum):
    EXPENSE_CLOTHING=1
    EXPENSE_FOOD=2
    EXPENSE_FUN=3
    EXPENSE_GIFTS=4
    EXPENSE_EAT_OUT=5
    EXPENSE_HOUSEHOLD=6
    EXPENSE_MISC=7
    EXPENSE_ROUTINE=8
    EXPENSE_TRAVEL=9
    INCOME_SAVINGS_ACC=10
    INCOME_CARD_ACC=11
    TRANSFER_SAVINGS_TO_CARD=12
    TRANSFER_SAVINGS_TO_STOCK=13
    REIMBURSEMENT_FOOD=14

payment_fields = {
    'G': CostType.EXPENSE_ROUTINE,
    'H': CostType.EXPENSE_FOOD,
    'I': CostType.EXPENSE_HOUSEHOLD,
    'J': CostType.EXPENSE_EAT_OUT,
    'K': CostType.EXPENSE_FUN,
    'L': CostType.EXPENSE_TRAVEL,
    'M': CostType.EXPENSE_CLOTHING,
    'N': CostType.EXPENSE_MISC
    #'': CostType.EXPENSE_GIFTS
}

cost_type_rules = {
    CostType.INCOME_SAVINGS_ACC: CostTypeRule('D', 'C'),
    CostType.INCOME_CARD_ACC: CostTypeRule('F', 'C'),
    CostType.TRANSFER_SAVINGS_TO_CARD: CostTypeRule('F', 'D'),
    CostType.TRANSFER_SAVINGS_TO_STOCK: CostTypeRule('E', 'D'),
    CostType.REIMBURSEMENT_FOOD: CostTypeRule('F', 'H')
}

for field in payment_fields.keys():
    cost_type_rules[payment_fields[field]] = CostTypeRule(field, 'F')

cost_type_rules[CostType.EXPENSE_ROUTINE] = CostTypeRule('G', 'D')

cost_type_translation_table = {
    "EMMAUS": CostType.EXPENSE_CLOTHING,
    "STADIUM DROTTNI": CostType.EXPENSE_CLOTHING,
    "HM SE0060": CostType.EXPENSE_CLOTHING,
    "HM SE0020": CostType.EXPENSE_CLOTHING,
    "HAIR & COSMETIC": CostType.EXPENSE_CLOTHING,
    "UNIQLO MALL OF": CostType.EXPENSE_CLOTHING,
    "NATURKOMPANIET /": CostType.EXPENSE_CLOTHING,
    "SPORTRINGOUTLET": CostType.EXPENSE_CLOTHING,
    "HEMKÖP DJURGÅRDS": CostType.EXPENSE_FOOD,
    "HEMKÖP SOLNA MAL": CostType.EXPENSE_FOOD,
    "ICA LAPPKARRSBER": CostType.EXPENSE_FOOD,
    "HEMKÖP CITY STOC": CostType.EXPENSE_FOOD,
    "MAX MALL OF SCAN": CostType.EXPENSE_FOOD,
    "ICA NARA SERGELS": CostType.EXPENSE_FOOD,
    "HIMALAYA LIVS": CostType.EXPENSE_FOOD,
    "PRESSBYRÅN 41082": CostType.EXPENSE_FOOD,
    "STOP 22": CostType.EXPENSE_FOOD,
    "PRESSBYRÅN 42501": CostType.EXPENSE_FOOD,
    "PRESSBYRÅN 40283": CostType.EXPENSE_FOOD,
    "COOP GUBBÄNGEN": CostType.EXPENSE_FOOD,
    "ICA NARA BERGSHA": CostType.EXPENSE_FOOD,
    "WILLYS FRIDHEMSP": CostType.EXPENSE_FOOD,
    "COOP KONSUM T-CE": CostType.EXPENSE_FOOD,
    "COOP KONSUM BERG": CostType.EXPENSE_FOOD,
    "TERMINALKIOSKEN": CostType.EXPENSE_FOOD,
    "NGROCERIES AB": CostType.EXPENSE_FOOD,
    "RAMEN KI MAMA": CostType.EXPENSE_EAT_OUT,
    "VETE-KATTEN AB": CostType.EXPENSE_EAT_OUT,
    "926446 RESTAURAN": CostType.EXPENSE_EAT_OUT,
    "RESTAURANG AND P": CostType.EXPENSE_EAT_OUT,
    "BURGER KING ODE": CostType.EXPENSE_EAT_OUT,
    "BAGERIET I POULT": CostType.EXPENSE_EAT_OUT,
    "SJ AB OMBORD": CostType.EXPENSE_EAT_OUT,
    "PROFESSORN RESTA": CostType.EXPENSE_EAT_OUT,
    "NON SOLO BAR ROR": CostType.EXPENSE_EAT_OUT,
    "R ASIA RESTAURAN": CostType.EXPENSE_EAT_OUT,
    "MAX STOCKHOLM VA": CostType.EXPENSE_EAT_OUT,
    "FRESH&FANCY": CostType.EXPENSE_EAT_OUT,
    "SARAVANAA INDISK": CostType.EXPENSE_EAT_OUT,
    "BEN & JERRY": CostType.EXPENSE_EAT_OUT,
    "iZ *Drop Coffee": CostType.EXPENSE_EAT_OUT,
    "JoeAndTheJuice": CostType.EXPENSE_EAT_OUT,
    "REGGEV HUMMUS": CostType.EXPENSE_EAT_OUT,
    "Rest Tegelbruket": CostType.EXPENSE_EAT_OUT,
    "LILLA KINA": CostType.EXPENSE_EAT_OUT,
    "PADELVERKET SPAN": CostType.EXPENSE_FUN,
    "SYSTEMBOLAGET": CostType.EXPENSE_FUN,
    "SYSTEMBOLAGET SO": CostType.EXPENSE_FUN,
    "SYSTEMBOLAGET RO": CostType.EXPENSE_FUN,
    "HELLSTEN MUSIK A": CostType.EXPENSE_FUN,
    "CLAS OHLSON": CostType.EXPENSE_HOUSEHOLD,
    "AB STORSTOCKHOL": CostType.EXPENSE_TRAVEL,
    "SL @STERMALMSTO": CostType.EXPENSE_TRAVEL,
    "SJ STOCKHOLM CEN": CostType.EXPENSE_TRAVEL,
    "Paynova AB (publ": CostType.EXPENSE_TRAVEL,
    "ROSLAGSBANAN T#": CostType.EXPENSE_TRAVEL,
    "CITYBANAN STATI": CostType.EXPENSE_TRAVEL,
    "SJ AB": CostType.EXPENSE_TRAVEL,
    "FOLKTANDVÅRD": CostType.EXPENSE_ROUTINE,
    "Bostadsförmedlin": CostType.EXPENSE_ROUTINE,
    "STIFT  STOCKHOLM": CostType.EXPENSE_ROUTINE,
    "LOOPIA AB": CostType.EXPENSE_ROUTINE,
    "Telia Mobile": CostType.EXPENSE_ROUTINE,
    "MUSESCORE PRO": CostType.EXPENSE_ROUTINE,
    "SLL Stockholms L": CostType.EXPENSE_ROUTINE,
    "Amazon Video*MV9": CostType.EXPENSE_ROUTINE,
    "CLAS OHLSON 218": CostType.EXPENSE_HOUSEHOLD,
    "APOTEK HJARTAT A": CostType.EXPENSE_HOUSEHOLD,
    "APOTEK HJARTAT M": CostType.EXPENSE_HOUSEHOLD,
    "APOTEKET C W SCH": CostType.EXPENSE_HOUSEHOLD,
    "APOTEKET SHOP": CostType.EXPENSE_HOUSEHOLD,
    "IKEA-KUNGENS KUR": CostType.EXPENSE_HOUSEHOLD,
    #"84319530719301": CostType.TRANSFER_SAVINGS_TO_CARD,
    "84319530717529": CostType.TRANSFER_SAVINGS_TO_CARD,
    "84319531718757": CostType.TRANSFER_SAVINGS_TO_STOCK,
    "@STRAS STATIONS": CostType.EXPENSE_MISC,
    "KONSTNARERNAS CE": CostType.EXPENSE_MISC,
    "SVEN HORNELL AB": CostType.EXPENSE_MISC,
    "WEBHALLEN": CostType.EXPENSE_MISC,
    "PANDURO HOBBY": CostType.EXPENSE_MISC,
    "PLANTAGEN FRESC": CostType.EXPENSE_MISC,
    "KREATIMA": CostType.EXPENSE_MISC,
    "457136731": CostType.EXPENSE_MISC, # Psykolog
    "LON": CostType.INCOME_SAVINGS_ACC,
    "GOD JUL": CostType.INCOME_SAVINGS_ACC,
    "1995082151350141": CostType.INCOME_CARD_ACC, # Sjukpenning
    "3504859": CostType.INCOME_SAVINGS_ACC, # Sjukpenning
    "ARVODE SSF": CostType.INCOME_SAVINGS_ACC,

    "+46738762178": CostType.EXPENSE_FOOD,
    "+46764134909": CostType.REIMBURSEMENT_FOOD
}

skipped_comments = [
    "84319530719301",
    "ATM T UNIVERSITE",
    "Utd TELIA",
    "Utd ASSA B"
]

# Read data file names
input_filenames = []
for i in range(1, len(sys.argv)-1):
    directory = sys.argv[i]
    for root, dirs, files in os.walk(directory, topdown=False):
        for name in files:
            if name.endswith(".xlsx"):
                input_filenames.append(os.path.join(root, name))

output_filename = sys.argv[-1]
wb_output = load_workbook(output_filename)

payments_summed = {}

# Collect payments for all sheets in the ouput file
for sheet_name in wb_output.sheetnames:
    r = 7
    sheet_out = wb_output.get_sheet_by_name(sheet_name)
    while sheet_out['B'+str(r)].value != None:
        comment=sheet_out['O'+str(r)].value

        # Read a specific cell
        payment = Payment(
            sheet_out['B'+str(r)].value,
            comment
        )

        # Sum cost
        for c in alphabet_uppercase[2:14:1] : 
            cost = sheet_out[c+str(r)].value
            if cost != None:
                # NOTE: Had to use -cost here in order to avoid sign problems, not sure why
                payment.set_cost(c, -cost)

        month_identifier = str.rsplit(payment.datetime, '-', 1)[0]
        if not month_identifier in payments_summed:
            payments_summed[month_identifier] = []
        payments_summed[month_identifier].append(payment)
        r = r+1

initial_balances = {
    "Transaktioner Privatkonto": 0,
    "Transaktioner e-sparkonto": 0,
    "Transaktioner Aktiekonto": 0
}

# Collect payments from input files
for filename in input_filenames:
    # Load xlsx workbook
    wb_input = load_workbook(filename)

    # Load active sheet or named sheet
    sheet_in = wb_input.active

    # Iterate over rows
    r = 8
    while sheet_in['A'+str(r)].value != None:
        r = r+1
        comment=sheet_in['E'+str(r)].value
        if comment == None:
            print("WARNING: Found payment comment that was None")
            continue
        if not comment in cost_type_translation_table:
            if not comment in skipped_comments:
                print("WARNING: Did not find comment: " + str(comment) + " in cost rules, please add it. Skipping for now...")
            continue

        # Read a specific cell
        payment = Payment(
            sheet_in['D'+str(r)].value,
            sheet_in['E'+str(r)].value,
        )

        cost_type = cost_type_translation_table[comment]
        cost_rule = cost_type_rules[cost_type]
        cost = abs(sheet_in['G'+str(r)].value)
        payment.set_cost(cost_rule.from_field, -cost)
        payment.set_cost(cost_rule.to_field, cost)

        month_identifier = str.rsplit(payment.datetime, '-', 1)[0]
        if not month_identifier in payments_summed:
            payments_summed[month_identifier] = []

        # Similar check. If similar, prioritize the payment from the output file,
        # since it can contain payments with multiple outputs
        similar_payment_found = False
        for payment_2 in payments_summed[month_identifier]:
            if payment.similar_to(payment_2):
                similar_payment_found = True
                break
        if similar_payment_found:
            continue

        #if not duplicate_found:
        payments_summed[month_identifier].append(payment)

    # Discern inital account balance
    initial_balance = 0
    i=9
    while sheet_in['H'+str(i)].value != None:
        initial_balance = sheet_in['H'+str(i)].value - sheet_in['G'+str(i)].value
        i+=1
    initial_balances[sheet_in["A1"].value] = initial_balance

print(initial_balances)

# Sort the months
months_to_iterate = []
for month_identifier in payments_summed.keys():
    months_to_iterate.append(month_identifier)
months_to_iterate.sort()

# Sort the payments per month
for month_identifier in payments_summed.keys():
    payments_summed[month_identifier].sort()

# Save initial buget
saved_data = {}
saved_data["initial_budget"] = []
saved_data["carryover"] = {}
saved_data["manual_changes"] = {}
saved_sheetnames = wb_output.sheetnames

if(months_to_iterate[0] in wb_output.sheetnames):
    first_sheet = wb_output.get_sheet_by_name(months_to_iterate[0])
    for c in alphabet_uppercase[6:14:1]: 
        saved_data["initial_budget"].append(first_sheet[c+"4"].value)

for month_identifier in months_to_iterate:
    if(month_identifier in wb_output.sheetnames):
        sheet = wb_output.get_sheet_by_name(month_identifier)
        saved_data["carryover"][month_identifier] = []
        saved_data["manual_changes"][month_identifier] = []

        footer_y = 1
        while sheet['B'+str(footer_y)].value != "Budgetdifferens, carryover":
            footer_y+=1

        for c in alphabet_uppercase[6:14:1]: 
            saved_data["carryover"][month_identifier].append(sheet[c+str(footer_y)].value)
            saved_data["manual_changes"][month_identifier].append(sheet[c+str(footer_y+1)].value)

#Refer to data in other sheets:
#='2020-12'!K23

print(saved_data)

# Clear output workbook
wb_output = Workbook()

for month_identifier in months_to_iterate:
    wb_output.create_sheet(month_identifier)
    sheet_out = wb_output.get_sheet_by_name(month_identifier)
    generate_header(sheet_out)
    put_payments(sheet_out, payments_summed[month_identifier])
    generate_footer(sheet_out, 8+len(payments_summed[month_identifier]))

    # Load saved data
    if(month_identifier in saved_sheetnames):
        if "initial budget" in saved_data.keys() and len(saved_data["initial budget"]) > 0:
            i=0
            for c in alphabet_uppercase[6:14:1]: 
                sheet_out[c+"4"].value = saved_data["initial_budget"][i]
                i+=1

        footer_y = 1
        while sheet_out['B'+str(footer_y)].value != "Budgetdifferens, carryover":
            footer_y+=1

        if "carryover" in saved_data.keys() and len(saved_data["carryover"]) > 0:
            i=0
            for c in alphabet_uppercase[6:14:1]: 
                sheet_out[c+str(footer_y)].value = saved_data["carryover"][month_identifier][i]
                sheet_out[c+str(footer_y+1)].value = saved_data["manual_changes"][month_identifier][i]
                i+=1

wb_output.remove_sheet(wb_output.active)

for month_index in range(0, len(months_to_iterate)-1):
    sheet_1 = wb_output.get_sheet_by_name(months_to_iterate[month_index])
    sheet_2 = wb_output.get_sheet_by_name(months_to_iterate[month_index+1])

    footer_y_1 = 1
    while sheet_1['B'+str(footer_y_1)].value != "Budget, ackumulerad nästa månad":
        footer_y_1+=1

    # Set budget reference to previous sheet
    for c in alphabet_uppercase[3:14:1] : 
        sheet_2[c+str(4)].value="='"+months_to_iterate[month_index]+"'!"+c+str(footer_y_1)

    # Set ingoing balance reference to previous sheet
    for c in alphabet_uppercase[3:5:1] : 
        sheet_2[c+str(5)].value="='"+months_to_iterate[month_index]+"'!"+c+str(footer_y_1-6)

wb_output.save(output_filename)
